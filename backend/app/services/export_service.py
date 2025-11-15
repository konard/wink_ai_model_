from io import BytesIO, StringIO
import csv
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

from ..models.script import Script, Scene


class ExportService:
    CATEGORY_LABELS_RU = {
        "violence": "Насилие",
        "gore": "Жестокость",
        "sex_act": "Сексуальный контент",
        "nudity": "Нагота",
        "profanity": "Ненормативная лексика",
        "drugs": "Наркотики",
        "child_risk": "Риск для детей"
    }

    @staticmethod
    def export_to_excel(
        script: Script,
        scenes: List[Scene],
        recommendations: List[Dict[str, Any]] = None
    ) -> BytesIO:
        wb = Workbook()

        wb.remove(wb.active)

        ws_overview = wb.create_sheet("Обзор")
        ExportService._create_overview_sheet(ws_overview, script)

        ws_scenes = wb.create_sheet("Сцены")
        ExportService._create_scenes_sheet(ws_scenes, scenes)

        if recommendations:
            ws_recs = wb.create_sheet("Рекомендации")
            ExportService._create_recommendations_sheet(ws_recs, recommendations)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _create_overview_sheet(ws, script: Script):
        header_font = Font(size=14, bold=True, color="1F4788")
        title_font = Font(size=16, bold=True)

        ws['A1'] = "Отчет по возрастному рейтингу"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        ws['A3'] = "Название"
        ws['B3'] = script.title
        ws['A4'] = "Рейтинг"
        ws['B4'] = script.predicted_rating or "—"
        ws['A5'] = "Всего сцен"
        ws['B5'] = script.total_scenes or 0
        ws['A6'] = "Дата создания"
        ws['B6'] = script.created_at.strftime("%d.%m.%Y %H:%M") if script.created_at else "—"

        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)

        if script.agg_scores:
            ws['A8'] = "Оценки по категориям"
            ws['A8'].font = header_font

            row = 9
            ws[f'A{row}'] = "Категория"
            ws[f'B{row}'] = "Оценка"
            ws[f'C{row}'] = "Оценка %"

            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

            for key, value in script.agg_scores.items():
                row += 1
                ws[f'A{row}'] = ExportService.CATEGORY_LABELS_RU.get(key, key)
                ws[f'B{row}'] = round(value, 3)
                ws[f'C{row}'] = f"{value*100:.1f}%"

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    @staticmethod
    def _create_scenes_sheet(ws, scenes: List[Scene]):
        headers = [
            "№ сцены", "Заголовок", "Насилие", "Жестокость",
            "Секс", "Нагота", "Мат", "Наркотики", "Риск детям"
        ]

        ws.append(headers)

        header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for scene in scenes:
            ws.append([
                scene.scene_id,
                scene.heading,
                round(scene.violence * 100, 1),
                round(scene.gore * 100, 1),
                round(scene.sex_act * 100, 1),
                round(scene.nudity * 100, 1),
                round(scene.profanity * 100, 1),
                round(scene.drugs * 100, 1),
                round(scene.child_risk * 100, 1)
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

    @staticmethod
    def _create_recommendations_sheet(ws, recommendations: List[Dict[str, Any]]):
        headers = ["№", "Описание", "Категория", "Сложность", "Влияние %", "Детали"]

        ws.append(headers)

        header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, rec in enumerate(recommendations, 1):
            specific_changes = "; ".join(rec.get('specific_changes', []))
            ws.append([
                idx,
                rec.get('description', ''),
                rec.get('category', ''),
                rec.get('difficulty', ''),
                round(rec.get('impact_score', 0) * 100, 1),
                specific_changes
            ])

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50

    @staticmethod
    def export_to_csv(
        script: Script,
        scenes: List[Scene]
    ) -> StringIO:
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow(['Отчет по возрастному рейтингу'])
        writer.writerow([])
        writer.writerow(['Название', script.title])
        writer.writerow(['Рейтинг', script.predicted_rating or '—'])
        writer.writerow(['Всего сцен', script.total_scenes or 0])
        writer.writerow(['Дата', script.created_at.strftime("%d.%m.%Y %H:%M") if script.created_at else '—'])
        writer.writerow([])

        if script.agg_scores:
            writer.writerow(['Оценки по категориям'])
            writer.writerow(['Категория', 'Оценка', 'Оценка %'])
            for key, value in script.agg_scores.items():
                label = ExportService.CATEGORY_LABELS_RU.get(key, key)
                writer.writerow([label, round(value, 3), f"{value*100:.1f}%"])
            writer.writerow([])

        writer.writerow(['Сцены'])
        writer.writerow([
            '№ сцены', 'Заголовок', 'Насилие %', 'Жестокость %',
            'Секс %', 'Нагота %', 'Мат %', 'Наркотики %', 'Риск детям %'
        ])

        for scene in scenes:
            writer.writerow([
                scene.scene_id,
                scene.heading,
                round(scene.violence * 100, 1),
                round(scene.gore * 100, 1),
                round(scene.sex_act * 100, 1),
                round(scene.nudity * 100, 1),
                round(scene.profanity * 100, 1),
                round(scene.drugs * 100, 1),
                round(scene.child_risk * 100, 1)
            ])

        output.seek(0)
        return output
